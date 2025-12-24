import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os

# Load data
df = pd.read_excel('graded_swim_data.xlsx')
df['Date'] = pd.to_datetime(df['Date'])

# Load standards for next standard calculation
with open('standards.json', 'r') as f:
    standards = json.load(f)

# Load goals (if exists)
if os.path.exists('goals.csv'):
    goals_df = pd.read_csv('goals.csv')
else:
    goals_df = pd.DataFrame(columns=['Event', 'Course', 'Goal_Time_Seconds', 'Goal_Standard', 'Notes'])

# Create workbook
wb = Workbook()
wb.remove(wb.active)

# ==================== SHEET 1: DASHBOARD ====================
dashboard = wb.create_sheet('Dashboard', 0)

# Title
dashboard['A1'] = 'SWIM PERFORMANCE DASHBOARD'
dashboard['A1'].font = Font(size=18, bold=True, color='FFFFFF')
dashboard['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
dashboard['A1'].alignment = Alignment(horizontal='center', vertical='center')
dashboard.merge_cells('A1:H1')
dashboard.row_dimensions[1].height = 30

# Key Metrics Section
dashboard['A3'] = 'KEY METRICS'
dashboard['A3'].font = Font(size=14, bold=True)
dashboard['A3'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
dashboard.merge_cells('A3:C3')

metrics = [
    ('Total Swims:', '=COUNTA(\'All Swims\'!A:A)-1'),
    ('AAA Times:', '=COUNTIF(\'All Swims\'!J:J,"AAA")'),
    ('AA Times:', '=COUNTIF(\'All Swims\'!J:J,"AA")'),
    ('A Times:', '=COUNTIF(\'All Swims\'!J:J,"A")'),
    ('Current Age:', f'={df["Age"].max()}'),
    ('First Meet:', f'=TEXT(MIN(\'All Swims\'!A:A),"MM/DD/YYYY")'),
    ('Latest Meet:', f'=TEXT(MAX(\'All Swims\'!A:A),"MM/DD/YYYY")'),
]

row = 4
for label, formula in metrics:
    dashboard[f'A{row}'] = label
    dashboard[f'A{row}'].font = Font(bold=True)
    dashboard[f'B{row}'] = formula
    dashboard[f'B{row}'].font = Font(size=12, color='0070C0')
    row += 1

# Standards Distribution
dashboard['E3'] = 'STANDARDS BREAKDOWN'
dashboard['E3'].font = Font(size=14, bold=True)
dashboard['E3'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
dashboard.merge_cells('E3:G3')

standards_list = ['AAAA', 'AAA', 'AA', 'A', 'BB', 'B', '<B']
row = 4
for std in standards_list:
    dashboard[f'E{row}'] = std
    dashboard[f'F{row}'] = f'=COUNTIF(\'All Swims\'!J:J,"{std}")'
    dashboard[f'E{row}'].font = Font(bold=True)
    dashboard[f'F{row}'].alignment = Alignment(horizontal='right')
    row += 1

# Personal Bests Lookup Section
dashboard['A13'] = 'PERSONAL BEST LOOKUP'
dashboard['A13'].font = Font(size=14, bold=True)
dashboard['A13'].fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
dashboard.merge_cells('A13:H13')

dashboard['A14'] = 'Select Event:'
dashboard['A14'].font = Font(bold=True)
dashboard['B14'] = '100 Free'
dashboard['B14'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
dashboard['B14'].font = Font(color='0000FF')

dashboard['D14'] = 'Select Course:'
dashboard['D14'].font = Font(bold=True)
dashboard['E14'] = 'Yards'
dashboard['E14'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
dashboard['E14'].font = Font(color='0000FF')

dashboard['A16'] = 'Best Time:'
dashboard['A16'].font = Font(bold=True)
dashboard['B16'] = '=IFERROR(VLOOKUP(B14&E14,\'Personal Bests\'!$H:$J,2,FALSE),"Not Found")'
dashboard['B16'].font = Font(size=12, bold=True, color='0070C0')

dashboard['A17'] = 'Standard:'
dashboard['A17'].font = Font(bold=True)
dashboard['B17'] = '=IFERROR(VLOOKUP(B14&E14,\'Personal Bests\'!$H:$J,3,FALSE),"Not Found")'
dashboard['B17'].font = Font(size=12, bold=True, color='0070C0')

dashboard['A18'] = 'Date:'
dashboard['A18'].font = Font(bold=True)
dashboard['B18'] = '=IFERROR(TEXT(VLOOKUP(B14&E14,\'Personal Bests\'!$H:$K,4,FALSE),"MM/DD/YYYY"),"Not Found")'
dashboard['B18'].font = Font(size=12, bold=True, color='0070C0')

# Instructions
dashboard['A20'] = 'INSTRUCTIONS:'
dashboard['A20'].font = Font(bold=True, size=11)
dashboard['A21'] = 'Click the dropdown arrows in B14 and E14 to select event and course'
dashboard['A21'].font = Font(italic=True)

# ==================== GOALS & PROGRESS SECTION ====================
goals_start_row = 23

dashboard[f'A{goals_start_row}'] = 'GOALS & PROGRESS'
dashboard[f'A{goals_start_row}'].font = Font(size=14, bold=True, color='FFFFFF')
dashboard[f'A{goals_start_row}'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
dashboard.merge_cells(f'A{goals_start_row}:H{goals_start_row}')
dashboard.row_dimensions[goals_start_row].height = 25

# Goals table header
goals_header_row = goals_start_row + 1
headers = ['Event', 'Course', 'Current Best', 'Goal Time', 'To Drop', '% Complete', 'Notes']
for col, header in enumerate(headers, 1):
    cell = dashboard.cell(goals_header_row, col, header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='C65911', end_color='C65911', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')

# Get personal bests for goal comparison
pb_data = df.copy()
pb_data = pb_data[pb_data['Standard'] != 'Unrated']
pb_data['Event'] = pb_data['Distance'].astype(str) + ' ' + pb_data['Stroke']
pb_dict = {}
for (event, course), group in pb_data.groupby(['Event', 'Course']):
    fastest = group.loc[group['Time_Seconds'].idxmin()]
    pb_dict[f"{event}_{course}"] = fastest['Time_Seconds']

# Add goals
goals_data_row = goals_header_row + 1
if len(goals_df) > 0:
    for _, goal in goals_df.iterrows():
        event = goal['Event']
        course = goal['Course']
        goal_time = goal['Goal_Time_Seconds']
        notes = goal.get('Notes', '')
        
        # Get current best
        current_best = pb_dict.get(f"{event}_{course}", None)
        
        dashboard.cell(goals_data_row, 1, event)
        dashboard.cell(goals_data_row, 2, course)
        
        if current_best:
            dashboard.cell(goals_data_row, 3, current_best)
            dashboard.cell(goals_data_row, 3).number_format = '0.00'
            dashboard.cell(goals_data_row, 4, goal_time)
            dashboard.cell(goals_data_row, 4).number_format = '0.00'
            
            # Calculate to drop
            to_drop = current_best - goal_time
            dashboard.cell(goals_data_row, 5, to_drop)
            dashboard.cell(goals_data_row, 5).number_format = '0.00'
            
            # Color code to_drop
            if to_drop > 0:
                dashboard.cell(goals_data_row, 5).font = Font(color='C00000', bold=True)  # Red - need to drop time
            else:
                dashboard.cell(goals_data_row, 5).font = Font(color='00B050', bold=True)  # Green - goal achieved!
            
            # Calculate % complete
            if to_drop >= 0:
                # Haven't reached goal yet
                progress = 0  # Starting point
            else:
                # Exceeded goal
                progress = 100
            
            dashboard.cell(goals_data_row, 6, progress)
            dashboard.cell(goals_data_row, 6).number_format = '0"%"'
        else:
            dashboard.cell(goals_data_row, 3, "No time yet")
            dashboard.cell(goals_data_row, 4, goal_time)
            dashboard.cell(goals_data_row, 4).number_format = '0.00'
        
        dashboard.cell(goals_data_row, 7, notes)
        dashboard.cell(goals_data_row, 7).font = Font(italic=True)
        
        goals_data_row += 1
else:
    # No goals yet - show message
    dashboard.cell(goals_data_row, 1, "No goals set yet! Edit goals.csv to add your target times.")
    dashboard.merge_cells(f'A{goals_data_row}:G{goals_data_row}')
    dashboard.cell(goals_data_row, 1).font = Font(italic=True, color='808080')
    goals_data_row += 1

# ==================== NEXT STANDARD SECTION ====================
next_std_start_row = goals_data_row + 2

dashboard[f'A{next_std_start_row}'] = 'NEXT STANDARD TARGETS'
dashboard[f'A{next_std_start_row}'].font = Font(size=14, bold=True, color='FFFFFF')
dashboard[f'A{next_std_start_row}'].fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
dashboard.merge_cells(f'A{next_std_start_row}:H{next_std_start_row}')
dashboard.row_dimensions[next_std_start_row].height = 25

# Next standard table header
next_std_header_row = next_std_start_row + 1
headers = ['Event', 'Course', 'Current Best', 'Current Std', 'Next Std', 'Target Time', 'To Drop']
for col, header in enumerate(headers, 1):
    cell = dashboard.cell(next_std_header_row, col, header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')

# Key events to track for next standard
key_events = [
    ('50 Free', 'Yards'),
    ('100 Free', 'Yards'),
    ('200 Free', 'Yards'),
    ('500 Free', 'Yards'),
    ('100 Back', 'Yards'),
    ('100 Breast', 'Yards'),
    ('100 Fly', 'Yards'),
    ('200 IM', 'Yards'),
]

def parse_time_standard(time_val):
    """Convert standard string to seconds"""
    if isinstance(time_val, (int, float)):
        return time_val
    try:
        if ':' in str(time_val):
            parts = str(time_val).split(':')
            return (float(parts[0]) * 60) + float(parts[1])
        return float(time_val)
    except:
        return 99999.0

def get_next_standard(current_std):
    """Get the next higher standard"""
    hierarchy = ['<B', 'B', 'BB', 'A', 'AA', 'AAA', 'AAAA']
    try:
        idx = hierarchy.index(current_std)
        if idx < len(hierarchy) - 1:
            return hierarchy[idx + 1]
    except:
        pass
    return None

next_std_data_row = next_std_header_row + 1
current_age = df['Age'].max()
current_date = df['Date'].max()

# Determine era and age group
if current_date < pd.Timestamp("2024-09-01"):
    era = "2021-2024"
else:
    era = "2024-2028"

if current_age <= 10:
    age_group = "10&U"
elif 11 <= current_age <= 12:
    age_group = "11-12"
elif 13 <= current_age <= 14:
    age_group = "13-14"
elif 15 <= current_age <= 16:
    age_group = "15-16"
else:
    age_group = "17-18"

for event, course in key_events:
    # Get current best
    current_best = pb_dict.get(f"{event}_{course}", None)
    
    if not current_best:
        continue
    
    # Get current standard from Personal Bests
    event_pbs = pb_data[(pb_data['Event'] == event) & (pb_data['Course'] == course)]
    if len(event_pbs) == 0:
        continue
    
    current_std = event_pbs.loc[event_pbs['Time_Seconds'].idxmin()]['Standard']
    
    # Get next standard
    next_std = get_next_standard(current_std)
    
    if not next_std:
        continue  # Already at AAAA
    
    # Look up target time for next standard
    try:
        course_key = 'SCY' if course == 'Yards' else 'LCM'
        event_standards = standards[era][age_group]['Male'][course_key][event]
        target_time_str = event_standards.get(next_std)
        
        if target_time_str:
            target_time = parse_time_standard(target_time_str)
            to_drop = current_best - target_time
            
            dashboard.cell(next_std_data_row, 1, event)
            dashboard.cell(next_std_data_row, 2, course)
            dashboard.cell(next_std_data_row, 3, current_best)
            dashboard.cell(next_std_data_row, 3).number_format = '0.00'
            dashboard.cell(next_std_data_row, 4, current_std)
            dashboard.cell(next_std_data_row, 5, next_std)
            
            # Color code next standard
            if next_std == 'AAAA':
                dashboard.cell(next_std_data_row, 5).fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
            elif next_std == 'AAA':
                dashboard.cell(next_std_data_row, 5).fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
            elif next_std == 'AA':
                dashboard.cell(next_std_data_row, 5).fill = PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid')
            
            dashboard.cell(next_std_data_row, 6, target_time)
            dashboard.cell(next_std_data_row, 6).number_format = '0.00'
            dashboard.cell(next_std_data_row, 7, to_drop)
            dashboard.cell(next_std_data_row, 7).number_format = '0.00'
            
            # Color code to_drop
            if to_drop <= 1.0:
                dashboard.cell(next_std_data_row, 7).fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
                dashboard.cell(next_std_data_row, 7).font = Font(bold=True, color='FFFFFF')
            elif to_drop <= 3.0:
                dashboard.cell(next_std_data_row, 7).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                dashboard.cell(next_std_data_row, 7).font = Font(bold=True)
            else:
                dashboard.cell(next_std_data_row, 7).font = Font(bold=True)
            
            next_std_data_row += 1
    except:
        continue  # Event not in standards

# Column widths
dashboard.column_dimensions['A'].width = 18
dashboard.column_dimensions['B'].width = 10
dashboard.column_dimensions['C'].width = 13
dashboard.column_dimensions['D'].width = 13
dashboard.column_dimensions['E'].width = 10
dashboard.column_dimensions['F'].width = 13
dashboard.column_dimensions['G'].width = 10
dashboard.column_dimensions['H'].width = 25

# ==================== SHEET 2: ALL SWIMS DATA ====================
all_swims = wb.create_sheet('All Swims', 1)

headers = ['Date', 'Age', 'Distance', 'Stroke', 'Round', 'Course', 'Time', 'Seconds', 'Meet', 'Standard']
for col, header in enumerate(headers, 1):
    cell = all_swims.cell(1, col, header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')

for row_idx, row in enumerate(df.itertuples(index=False), 2):
    all_swims.cell(row_idx, 1, row.Date)
    all_swims.cell(row_idx, 2, row.Age)
    all_swims.cell(row_idx, 3, row.Distance)
    all_swims.cell(row_idx, 4, row.Stroke)
    all_swims.cell(row_idx, 5, row.Round)
    all_swims.cell(row_idx, 6, row.Course)
    all_swims.cell(row_idx, 7, row.Finals)
    all_swims.cell(row_idx, 8, row.Time_Seconds)
    all_swims.cell(row_idx, 9, row.Meet)
    all_swims.cell(row_idx, 10, row.Standard)

all_swims.column_dimensions['A'].width = 12
all_swims.column_dimensions['I'].width = 35
all_swims.column_dimensions['J'].width = 10

# ==================== SHEET 3: PERSONAL BESTS ====================
pb_sheet = wb.create_sheet('Personal Bests', 2)

headers = ['EVENT', 'COURSE', 'BEST TIME', 'SECONDS', 'STANDARD', 'DATE', 'AGE', 'LOOKUP_KEY', 'TIME', 'STD', 'DATE2']
for col, header in enumerate(headers, 1):
    pb_sheet.cell(1, col, header)
    pb_sheet.cell(1, col).font = Font(bold=True, color='FFFFFF')
    pb_sheet.cell(1, col).fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')

pb_summary = []
for (event, course), group in pb_data.groupby(['Event', 'Course']):
    fastest = group.loc[group['Time_Seconds'].idxmin()]
    pb_summary.append({
        'Event': event,
        'Course': course,
        'Time': fastest['Finals'],
        'Seconds': fastest['Time_Seconds'],
        'Standard': fastest['Standard'],
        'Date': fastest['Date'],
        'Age': fastest['Age']
    })

pb_df = pd.DataFrame(pb_summary).sort_values(['Event', 'Course'])
unique_events = sorted(pb_df['Event'].unique())

row = 2
for _, pb in pb_df.iterrows():
    pb_sheet.cell(row, 1, pb['Event'])
    pb_sheet.cell(row, 2, pb['Course'])
    pb_sheet.cell(row, 3, pb['Time'])
    pb_sheet.cell(row, 4, pb['Seconds'])
    pb_sheet.cell(row, 5, pb['Standard'])
    pb_sheet.cell(row, 6, pb['Date'])
    pb_sheet.cell(row, 7, pb['Age'])
    pb_sheet.cell(row, 8, f"=A{row}&B{row}")
    pb_sheet.cell(row, 9, pb['Time'])
    pb_sheet.cell(row, 10, pb['Standard'])
    pb_sheet.cell(row, 11, pb['Date'])
    
    std_cell = pb_sheet.cell(row, 5)
    if pb['Standard'] == 'AAAA':
        std_cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
    elif pb['Standard'] == 'AAA':
        std_cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
    elif pb['Standard'] == 'AA':
        std_cell.fill = PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid')
    elif pb['Standard'] == 'A':
        std_cell.fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
    
    row += 1

pb_sheet.column_dimensions['A'].width = 15
pb_sheet.column_dimensions['B'].width = 10
pb_sheet.column_dimensions['C'].width = 12
pb_sheet.column_dimensions['F'].width = 12
pb_sheet.column_dimensions['H'].hidden = True
pb_sheet.column_dimensions['I'].hidden = True
pb_sheet.column_dimensions['J'].hidden = True
pb_sheet.column_dimensions['K'].hidden = True

# ==================== SHEET 4: TIME PROGRESSIONS ====================
progress_sheet = wb.create_sheet('Time Progressions', 3)

key_events_prog = [
    ('50 Free', 'Yards'),
    ('100 Free', 'Yards'),
    ('200 Free', 'Yards'),
    ('500 Free', 'Yards'),
    ('100 Back', 'Yards'),
    ('100 Breast', 'Yards'),
    ('100 Fly', 'Yards'),
    ('200 IM', 'Yards'),
]

row_start = 1
for event_name, course in key_events_prog:
    event_data = df[(df['Distance'].astype(str) + ' ' + df['Stroke'] == event_name) & 
                    (df['Course'] == course)].copy()
    
    if len(event_data) == 0:
        continue
    
    event_data = event_data.sort_values('Date')
    
    progress_sheet.cell(row_start, 1, f'{event_name} - {course}')
    progress_sheet.cell(row_start, 1).font = Font(size=12, bold=True)
    progress_sheet.cell(row_start, 1).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    
    progress_sheet.cell(row_start + 1, 1, 'Date')
    progress_sheet.cell(row_start + 1, 2, 'Time (sec)')
    progress_sheet.cell(row_start + 1, 3, 'Standard')
    
    for col in range(1, 4):
        progress_sheet.cell(row_start + 1, col).font = Font(bold=True)
    
    data_row = row_start + 2
    for _, swim in event_data.iterrows():
        progress_sheet.cell(data_row, 1, swim['Date'])
        progress_sheet.cell(data_row, 2, swim['Time_Seconds'])
        progress_sheet.cell(data_row, 3, swim['Standard'])
        data_row += 1
    
    chart = LineChart()
    chart.title = f'{event_name} Progression'
    chart.y_axis.title = 'Time (seconds)'
    chart.x_axis.title = 'Date'
    
    data = Reference(progress_sheet, min_col=2, min_row=row_start + 1, max_row=data_row - 1)
    dates = Reference(progress_sheet, min_col=1, min_row=row_start + 2, max_row=data_row - 1)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(dates)
    
    chart.width = 15
    chart.height = 10
    progress_sheet.add_chart(chart, f'E{row_start}')
    
    row_start = data_row + 3

progress_sheet.column_dimensions['A'].width = 12
progress_sheet.column_dimensions['B'].width = 12
progress_sheet.column_dimensions['C'].width = 12

# ==================== SHEET 5: DROPDOWN LISTS ====================
lists_sheet = wb.create_sheet('DropdownLists', 4)
lists_sheet.sheet_state = 'hidden'

lists_sheet['A1'] = 'EVENTS'
lists_sheet['A1'].font = Font(bold=True)
for idx, event in enumerate(unique_events, 2):
    lists_sheet[f'A{idx}'] = event

lists_sheet['B1'] = 'COURSES'
lists_sheet['B1'].font = Font(bold=True)
lists_sheet['B2'] = 'Yards'
lists_sheet['B3'] = 'LCM'

# ==================== ADD DATA VALIDATION (DROPDOWNS) ====================
event_dropdown = DataValidation(
    type="list",
    formula1=f"=DropdownLists!$A$2:$A${len(unique_events)+1}",
    allow_blank=False
)
event_dropdown.error = 'Please select from the dropdown list'
event_dropdown.errorTitle = 'Invalid Event'
dashboard.add_data_validation(event_dropdown)
event_dropdown.add(dashboard['B14'])

course_dropdown = DataValidation(
    type="list",
    formula1="=DropdownLists!$B$2:$B$3",
    allow_blank=False
)
course_dropdown.error = 'Please select either Yards or LCM'
course_dropdown.errorTitle = 'Invalid Course'
dashboard.add_data_validation(course_dropdown)
course_dropdown.add(dashboard['E14'])

wb.save('Swim_Dashboard.xlsx')
print('✅ Dashboard created: Swim_Dashboard.xlsx')
print('\n🎯 NEW FEATURES:')
print('  ✅ GOALS & PROGRESS - Track your target times')
print('  ✅ NEXT STANDARD TARGETS - See what you need for the next level')
print('  ✅ Color-coded progress indicators')
print(f'  ✅ Tracking {len(goals_df)} goals')
print('\nSheets created:')
print('  📊 Dashboard - Now with Goals & Next Standard sections!')
print('  📋 All Swims - Complete swim history')
print('  🏆 Personal Bests - Best times by event')
print('  📈 Time Progressions - Charts for key events')
print('  🔽 DropdownLists - Hidden sheet with dropdown values')
print('\n💡 TIP: Edit goals.csv to add your season goals!')
